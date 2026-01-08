"""
Export de rapports analytics (Excel, PDF, CSV) pour le Mode Gestionnaire
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from datetime import datetime, date, timedelta
from typing import Optional, Literal
import io
import csv
import random

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference

# PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from app.dependencies.db_deps import get_session
from app.models_structure import Lit, Service, UniteFonctionnelle


router = APIRouter(prefix="/api/analytics/export", tags=["Analytics Export"])


def get_kpi_data(session: Session, eg_id: int, period: str):
    """Récupère les données KPIs pour export (simulations MVP)"""
    # Simulation des KPIs (même logique que dans analytics.py)
    total_lits = session.exec(
        select(Lit).where(Lit.eg_id == eg_id)
    ).all()
    
    nb_lits_total = len(total_lits)
    nb_lits_occupes = int(nb_lits_total * random.uniform(0.65, 0.85))
    taux_occupation = (nb_lits_occupes / nb_lits_total * 100) if nb_lits_total > 0 else 0
    
    return {
        "taux_occupation": round(taux_occupation, 1),
        "dms": round(random.uniform(4.5, 7.2), 1),
        "taux_rotation": round(random.uniform(35, 55), 1),
        "lits_disponibles": nb_lits_total - nb_lits_occupes,
        "taux_ouverture": round(random.uniform(85, 98), 1),
        "nb_lits_total": nb_lits_total,
        "nb_lits_occupes": nb_lits_occupes,
    }


def get_capacity_data(session: Session, eg_id: int):
    """Récupère les données de capacité par service"""
    services = session.exec(
        select(Service).where(Service.eg_id == eg_id)
    ).all()
    
    data = []
    for service in services:
        lits = session.exec(
            select(Lit).where(Lit.service_id == service.id)
        ).all()
        
        nb_lits = len(lits)
        if nb_lits > 0:
            occupation = random.uniform(65, 95)
            data.append({
                "service": service.nom or f"Service {service.code}",
                "nb_lits": nb_lits,
                "taux_occupation": round(occupation, 1),
                "lits_occupes": int(nb_lits * occupation / 100),
                "lits_disponibles": nb_lits - int(nb_lits * occupation / 100),
            })
    
    return data


@router.get("/excel")
async def export_excel(
    eg_id: int = Query(1, description="ID de l'Entité Géographique"),
    period: str = Query("30d", description="Période (7d, 30d, 1y)"),
    session: Session = Depends(get_session)
):
    """Export Excel avec tableaux et graphiques"""
    
    # Créer le workbook
    wb = Workbook()
    
    # === FEUILLE 1: KPIs ===
    ws_kpis = wb.active
    ws_kpis.title = "KPIs"
    
    # En-tête
    ws_kpis['A1'] = "Rapport Analytics - Mode Gestionnaire"
    ws_kpis['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_kpis['A1'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_kpis['A1'].alignment = Alignment(horizontal="center")
    ws_kpis.merge_cells('A1:E1')
    
    ws_kpis['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_kpis['A2'].alignment = Alignment(horizontal="center")
    ws_kpis.merge_cells('A2:E2')
    
    # Récupérer les KPIs
    kpis = get_kpi_data(session, eg_id, period)
    
    # Tableau KPIs
    ws_kpis['A4'] = "Indicateur"
    ws_kpis['B4'] = "Valeur"
    ws_kpis['C4'] = "Unité"
    
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    for cell in ['A4', 'B4', 'C4']:
        ws_kpis[cell].font = Font(bold=True)
        ws_kpis[cell].fill = header_fill
    
    kpi_rows = [
        ("Taux d'occupation", kpis["taux_occupation"], "%"),
        ("Durée Moyenne de Séjour", kpis["dms"], "jours"),
        ("Taux de rotation", kpis["taux_rotation"], "%"),
        ("Lits disponibles", kpis["lits_disponibles"], "lits"),
        ("Taux d'ouverture", kpis["taux_ouverture"], "%"),
        ("Lits totaux", kpis["nb_lits_total"], "lits"),
        ("Lits occupés", kpis["nb_lits_occupes"], "lits"),
    ]
    
    for idx, (label, value, unit) in enumerate(kpi_rows, start=5):
        ws_kpis[f'A{idx}'] = label
        ws_kpis[f'B{idx}'] = value
        ws_kpis[f'C{idx}'] = unit
    
    # Ajuster les colonnes
    ws_kpis.column_dimensions['A'].width = 30
    ws_kpis.column_dimensions['B'].width = 15
    ws_kpis.column_dimensions['C'].width = 10
    
    # === FEUILLE 2: Capacité par Service ===
    ws_capacity = wb.create_sheet("Capacité Services")
    
    ws_capacity['A1'] = "Capacité par Service"
    ws_capacity['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_capacity['A1'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws_capacity['A1'].alignment = Alignment(horizontal="center")
    ws_capacity.merge_cells('A1:E1')
    
    # Headers
    headers = ["Service", "Lits totaux", "Lits occupés", "Lits disponibles", "Taux occupation (%)"]
    for idx, header in enumerate(headers, start=1):
        cell = ws_capacity.cell(row=3, column=idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
    
    # Données
    capacity_data = get_capacity_data(session, eg_id)
    for row_idx, service_data in enumerate(capacity_data, start=4):
        ws_capacity.cell(row=row_idx, column=1, value=service_data["service"])
        ws_capacity.cell(row=row_idx, column=2, value=service_data["nb_lits"])
        ws_capacity.cell(row=row_idx, column=3, value=service_data["lits_occupes"])
        ws_capacity.cell(row=row_idx, column=4, value=service_data["lits_disponibles"])
        ws_capacity.cell(row=row_idx, column=5, value=service_data["taux_occupation"])
        
        # Coloration selon taux
        taux = service_data["taux_occupation"]
        if taux > 95:
            fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        elif taux > 80:
            fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        
        ws_capacity.cell(row=row_idx, column=5).fill = fill
    
    # Ajuster colonnes
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_capacity.column_dimensions[col].width = 20
    
    # Ajouter un graphique
    if len(capacity_data) > 0:
        chart = BarChart()
        chart.title = "Capacité par service"
        chart.x_axis.title = "Services"
        chart.y_axis.title = "Nombre de lits"
        
        data = Reference(ws_capacity, min_col=2, min_row=3, max_row=3+len(capacity_data), max_col=4)
        cats = Reference(ws_capacity, min_col=1, min_row=4, max_row=3+len(capacity_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        
        ws_capacity.add_chart(chart, "G3")
    
    # Sauvegarder dans un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"rapport_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/pdf")
async def export_pdf(
    eg_id: int = Query(1, description="ID de l'Entité Géographique"),
    period: str = Query("30d", description="Période (7d, 30d, 1y)"),
    session: Session = Depends(get_session)
):
    """Export PDF formaté avec tableaux"""
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Style personnalisé pour le titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#3B82F6'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # Titre
    elements.append(Paragraph("📊 Rapport Analytics", title_style))
    elements.append(Paragraph(f"Mode Gestionnaire - Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", 
                              styles['Normal']))
    elements.append(Spacer(1, 1*cm))
    
    # Section KPIs
    elements.append(Paragraph("Indicateurs Clés de Performance", styles['Heading2']))
    elements.append(Spacer(1, 0.5*cm))
    
    kpis = get_kpi_data(session, eg_id, period)
    
    kpi_data = [
        ['Indicateur', 'Valeur', 'Unité'],
        ["Taux d'occupation", f"{kpis['taux_occupation']}", "%"],
        ["Durée Moyenne de Séjour", f"{kpis['dms']}", "jours"],
        ["Taux de rotation", f"{kpis['taux_rotation']}", "%"],
        ["Lits disponibles", f"{kpis['lits_disponibles']}", "lits"],
        ["Taux d'ouverture", f"{kpis['taux_ouverture']}", "%"],
    ]
    
    kpi_table = Table(kpi_data, colWidths=[10*cm, 4*cm, 3*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(kpi_table)
    elements.append(Spacer(1, 1*cm))
    
    # Section Capacité
    elements.append(Paragraph("Capacité par Service", styles['Heading2']))
    elements.append(Spacer(1, 0.5*cm))
    
    capacity_data = get_capacity_data(session, eg_id)
    
    capacity_table_data = [
        ['Service', 'Lits totaux', 'Occupés', 'Disponibles', 'Taux (%)']
    ]
    
    for service in capacity_data[:10]:  # Limiter à 10 services pour le PDF
        capacity_table_data.append([
            service['service'],
            str(service['nb_lits']),
            str(service['lits_occupes']),
            str(service['lits_disponibles']),
            f"{service['taux_occupation']}%"
        ])
    
    capacity_table = Table(capacity_table_data, colWidths=[6*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    capacity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(capacity_table)
    
    # Note de bas de page
    elements.append(Spacer(1, 2*cm))
    elements.append(Paragraph(
        "Rapport généré automatiquement par MedData Bridge - Mode Gestionnaire",
        styles['Normal']
    ))
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rapport_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/csv")
async def export_csv(
    eg_id: int = Query(1, description="ID de l'Entité Géographique"),
    period: str = Query("30d", description="Période (7d, 30d, 1y)"),
    export_type: Literal["kpis", "capacity"] = Query("capacity", description="Type de données à exporter"),
    session: Session = Depends(get_session)
):
    """Export CSV des données brutes"""
    
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    
    if export_type == "kpis":
        # Export KPIs
        writer.writerow(["Indicateur", "Valeur", "Unité", "Date export"])
        
        kpis = get_kpi_data(session, eg_id, period)
        date_export = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        writer.writerow(["Taux d'occupation", kpis["taux_occupation"], "%", date_export])
        writer.writerow(["Durée Moyenne de Séjour", kpis["dms"], "jours", date_export])
        writer.writerow(["Taux de rotation", kpis["taux_rotation"], "%", date_export])
        writer.writerow(["Lits disponibles", kpis["lits_disponibles"], "lits", date_export])
        writer.writerow(["Taux d'ouverture", kpis["taux_ouverture"], "%", date_export])
        writer.writerow(["Lits totaux", kpis["nb_lits_total"], "lits", date_export])
        writer.writerow(["Lits occupés", kpis["nb_lits_occupes"], "lits", date_export])
        
    else:  # capacity
        # Export capacité par service
        writer.writerow(["Service", "Lits totaux", "Lits occupés", "Lits disponibles", "Taux occupation (%)", "Date export"])
        
        capacity_data = get_capacity_data(session, eg_id)
        date_export = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        for service in capacity_data:
            writer.writerow([
                service["service"],
                service["nb_lits"],
                service["lits_occupes"],
                service["lits_disponibles"],
                service["taux_occupation"],
                date_export
            ])
    
    buffer.seek(0)
    
    filename = f"export_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
