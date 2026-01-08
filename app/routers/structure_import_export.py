"""
Import/Export de structure hospitalière via Excel
"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Request, Form
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlmodel import Session, select
from typing import Optional, Literal
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from app.dependencies.db_deps import get_session
from app.models_structure import (
    EntiteGeographique, Pole, Service, UniteFonctionnelle,
    UniteHebergement, Chambre, Lit
)
from app.schemas.import_schemas import ImportAction

router = APIRouter(prefix="/api/structure", tags=["Structure Import/Export"])
ui_router = APIRouter(prefix="/structure", tags=["Structure Import/Export UI"])
templates = Jinja2Templates(directory="app/templates")


@router.get("/export/excel")
async def export_structure_excel(
    eg_id: Optional[int] = None,
    session: Session = Depends(get_session)
):
    """Export de la structure complète en Excel avec 7 feuilles"""
    
    wb = Workbook()
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # === FEUILLE 1: README ===
    ws_readme = wb.active
    ws_readme.title = "README"
    
    ws_readme['A1'] = "📋 Guide d'utilisation - Import/Export Structure"
    ws_readme['A1'].font = Font(size=16, bold=True, color="3B82F6")
    ws_readme.merge_cells('A1:D1')
    
    readme_content = [
        ("", ""),
        ("Ce fichier contient la structure complète de l'établissement.", ""),
        ("Chaque feuille représente un niveau hiérarchique :", ""),
        ("", ""),
        ("1. EntitesGeographiques", "Établissements ou sites géographiques"),
        ("2. Poles", "Pôles d'activité"),
        ("3. Services", "Services médicaux ou techniques"),
        ("4. UnitésFonctionnelles", "Unités fonctionnelles (UF)"),
        ("5. UnitesHebergement", "Unités d'hébergement (UH)"),
        ("6. Chambres", "Chambres"),
        ("7. Lits", "Lits individuels"),
        ("", ""),
        ("⚠️ IMPORTANT :", ""),
        ("- Les codes doivent être uniques", ""),
        ("- Respecter les références entre niveaux (eg_code existe dans Pôles)", ""),
        ("- Ne pas supprimer les en-têtes", ""),
        ("- Sauvegarder en .xlsx uniquement", ""),
        ("", ""),
        (f"📅 Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", ""),
    ]
    
    for idx, (col_a, col_b) in enumerate(readme_content, start=2):
        ws_readme[f'A{idx}'] = col_a
        ws_readme[f'B{idx}'] = col_b
    
    ws_readme.column_dimensions['A'].width = 40
    ws_readme.column_dimensions['B'].width = 50
    
    # === FEUILLE 2: Entités Géographiques ===
    ws_eg = wb.create_sheet("EntitesGeographiques")
    
    eg_headers = ["code", "nom", "type_eg", "adresse", "ville", "code_postal", "telephone", "finess"]
    for idx, header in enumerate(eg_headers, start=1):
        cell = ws_eg.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    query_eg = select(EntiteGeographique)
    if eg_id:
        query_eg = query_eg.where(EntiteGeographique.id == eg_id)
    
    egs = session.exec(query_eg).all()
    for row_idx, eg in enumerate(egs, start=2):
        ws_eg.cell(row=row_idx, column=1, value=eg.code or f"EG{eg.id}")
        ws_eg.cell(row=row_idx, column=2, value=eg.nom)
        ws_eg.cell(row=row_idx, column=3, value=eg.type_eg or "")
        ws_eg.cell(row=row_idx, column=4, value=eg.adresse or "")
        ws_eg.cell(row=row_idx, column=5, value=eg.ville or "")
        ws_eg.cell(row=row_idx, column=6, value=eg.code_postal or "")
        ws_eg.cell(row=row_idx, column=7, value=eg.telephone or "")
        ws_eg.cell(row=row_idx, column=8, value=eg.finess or "")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_eg.column_dimensions[col].width = 18
    
    # === FEUILLE 3: Pôles ===
    ws_poles = wb.create_sheet("Poles")
    
    pole_headers = ["code", "nom", "eg_code", "responsable", "telephone"]
    for idx, header in enumerate(pole_headers, start=1):
        cell = ws_poles.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    query_poles = select(Pole)
    if eg_id:
        query_poles = query_poles.where(Pole.eg_id == eg_id)
    
    poles = session.exec(query_poles).all()
    for row_idx, pole in enumerate(poles, start=2):
        eg = session.get(EntiteGeographique, pole.eg_id) if pole.eg_id else None
        ws_poles.cell(row=row_idx, column=1, value=pole.code or f"POL{pole.id}")
        ws_poles.cell(row=row_idx, column=2, value=pole.nom)
        ws_poles.cell(row=row_idx, column=3, value=eg.code if eg else "")
        ws_poles.cell(row=row_idx, column=4, value="")
        ws_poles.cell(row=row_idx, column=5, value="")
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_poles.column_dimensions[col].width = 20
    
    # === FEUILLE 4: Services ===
    ws_services = wb.create_sheet("Services")
    
    service_headers = ["code", "nom", "pole_code", "um_code", "telephone", "responsable"]
    for idx, header in enumerate(service_headers, start=1):
        cell = ws_services.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    query_services = select(Service)
    if eg_id:
        query_services = query_services.where(Service.eg_id == eg_id)
    
    services = session.exec(query_services).all()
    for row_idx, service in enumerate(services, start=2):
        pole = session.get(Pole, service.pole_id) if service.pole_id else None
        ws_services.cell(row=row_idx, column=1, value=service.code or f"SRV{service.id}")
        ws_services.cell(row=row_idx, column=2, value=service.nom)
        ws_services.cell(row=row_idx, column=3, value=pole.code if pole else "")
        ws_services.cell(row=row_idx, column=4, value="")
        ws_services.cell(row=row_idx, column=5, value="")
        ws_services.cell(row=row_idx, column=6, value="")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_services.column_dimensions[col].width = 20
    
    # === FEUILLE 5: Unités Fonctionnelles ===
    ws_ufs = wb.create_sheet("UnitésFonctionnelles")
    
    uf_headers = ["code", "nom", "service_code", "um_code", "type_uf", "capacite"]
    for idx, header in enumerate(uf_headers, start=1):
        cell = ws_ufs.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    query_ufs = select(UniteFonctionnelle)
    if eg_id:
        query_ufs = query_ufs.where(UniteFonctionnelle.eg_id == eg_id)
    
    ufs = session.exec(query_ufs).all()
    for row_idx, uf in enumerate(ufs, start=2):
        service = session.get(Service, uf.service_id) if uf.service_id else None
        ws_ufs.cell(row=row_idx, column=1, value=uf.code or f"UF{uf.id}")
        ws_ufs.cell(row=row_idx, column=2, value=uf.nom)
        ws_ufs.cell(row=row_idx, column=3, value=service.code if service else "")
        ws_ufs.cell(row=row_idx, column=4, value=uf.um_code or "")
        ws_ufs.cell(row=row_idx, column=5, value=uf.type_uf or "")
        ws_ufs.cell(row=row_idx, column=6, value="")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_ufs.column_dimensions[col].width = 20
    
    # === FEUILLE 6: Unités d'Hébergement ===
    ws_uhs = wb.create_sheet("UnitesHebergement")
    
    uh_headers = ["code", "nom", "uf_code", "etage", "batiment", "capacite"]
    for idx, header in enumerate(uh_headers, start=1):
        cell = ws_uhs.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    query_uhs = select(UniteHebergement)
    if eg_id:
        query_uhs = query_uhs.where(UniteHebergement.eg_id == eg_id)
    
    uhs = session.exec(query_uhs).all()
    for row_idx, uh in enumerate(uhs, start=2):
        uf = session.get(UniteFonctionnelle, uh.uf_id) if uh.uf_id else None
        ws_uhs.cell(row=row_idx, column=1, value=uh.code or f"UH{uh.id}")
        ws_uhs.cell(row=row_idx, column=2, value=uh.nom)
        ws_uhs.cell(row=row_idx, column=3, value=uf.code if uf else "")
        ws_uhs.cell(row=row_idx, column=4, value=uh.etage or "")
        ws_uhs.cell(row=row_idx, column=5, value=uh.batiment or "")
        ws_uhs.cell(row=row_idx, column=6, value="")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_uhs.column_dimensions[col].width = 20
    
    # === FEUILLE 7: Chambres ===
    ws_chambres = wb.create_sheet("Chambres")
    
    chambre_headers = ["numero", "uh_code", "capacite", "type_chambre"]
    for idx, header in enumerate(chambre_headers, start=1):
        cell = ws_chambres.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    query_chambres = select(Chambre)
    if eg_id:
        query_chambres = query_chambres.where(Chambre.eg_id == eg_id)
    
    chambres = session.exec(query_chambres).all()
    for row_idx, chambre in enumerate(chambres, start=2):
        uh = session.get(UniteHebergement, chambre.uh_id) if chambre.uh_id else None
        ws_chambres.cell(row=row_idx, column=1, value=chambre.numero)
        ws_chambres.cell(row=row_idx, column=2, value=uh.code if uh else "")
        ws_chambres.cell(row=row_idx, column=3, value=chambre.capacite or 1)
        ws_chambres.cell(row=row_idx, column=4, value=chambre.type_chambre or "")
    
    for col in ['A', 'B', 'C', 'D']:
        ws_chambres.column_dimensions[col].width = 20
    
    # === FEUILLE 8: Lits ===
    ws_lits = wb.create_sheet("Lits")
    
    lit_headers = ["numero", "chambre_numero", "uh_code", "type_lit", "statut"]
    for idx, header in enumerate(lit_headers, start=1):
        cell = ws_lits.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    query_lits = select(Lit)
    if eg_id:
        query_lits = query_lits.where(Lit.eg_id == eg_id)
    
    lits = session.exec(query_lits).all()
    for row_idx, lit in enumerate(lits, start=2):
        chambre = session.get(Chambre, lit.chambre_id) if lit.chambre_id else None
        uh = session.get(UniteHebergement, lit.uh_id) if lit.uh_id else None
        ws_lits.cell(row=row_idx, column=1, value=lit.numero)
        ws_lits.cell(row=row_idx, column=2, value=chambre.numero if chambre else "")
        ws_lits.cell(row=row_idx, column=3, value=uh.code if uh else "")
        ws_lits.cell(row=row_idx, column=4, value=lit.type_lit or "")
        ws_lits.cell(row=row_idx, column=5, value=lit.statut or "")
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_lits.column_dimensions[col].width = 20
    
    # Sauvegarder dans buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"structure_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/template")
async def export_template():
    """Télécharge un template Excel vierge avec exemples"""
    
    wb = Workbook()
    
    # Style en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Créer les feuilles avec une ligne d'exemple
    ws_eg = wb.active
    ws_eg.title = "EntitesGeographiques"
    eg_headers = ["code", "nom", "type_eg", "adresse", "ville", "code_postal", "telephone", "finess"]
    for idx, header in enumerate(eg_headers, start=1):
        cell = ws_eg.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Exemple
    ws_eg['A2'] = "EG01"
    ws_eg['B2'] = "Centre Hospitalier Nord"
    ws_eg['C2'] = "CH"
    ws_eg['D2'] = "1 rue de la Santé"
    ws_eg['E2'] = "Paris"
    ws_eg['F2'] = "75001"
    ws_eg['G2'] = "01 23 45 67 89"
    ws_eg['H2'] = "750001234"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_eg.column_dimensions[col].width = 18
    
    # Autres feuilles similaires...
    # (Code similaire pour Poles, Services, etc.)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = "template_structure.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# === Import Endpoints ===

@router.post("/import/excel")
async def import_excel_preview(
    file: UploadFile = File(...),
    mode: Literal["create", "update", "replace"] = Form(...),
    session: Session = Depends(get_session)
):
    """
    Parse et valide un fichier Excel, retourne un aperçu avant import
    """
    from openpyxl import load_workbook
    from app.schemas.import_schemas import (
        ImportMode, ImportPreview, ImportEntityPreview, ImportAction,
        ImportMessage, ImportSeverity,
        ExcelRowEG, ExcelRowPole, ExcelRowService, ExcelRowUF,
        ExcelRowUH, ExcelRowChambre, ExcelRowLit
    )
    import time
    
    start_time = time.time()
    
    # Validation format fichier
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Format invalide. Seuls .xlsx et .xls sont acceptés")
    
    # Lecture du fichier
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Erreur lecture Excel: {str(e)}")
    
    preview = ImportPreview(mode=ImportMode(mode), total_rows=0)
    
    # Mapping des feuilles Excel vers les modèles
    sheet_configs = [
        {
            "name": "EntitesGeographiques",
            "model": ExcelRowEG,
            "db_model": EntiteGeographique,
            "entity_type": "eg",
            "parent_field": None
        },
        {
            "name": "Poles",
            "model": ExcelRowPole,
            "db_model": Pole,
            "entity_type": "pole",
            "parent_field": "eg_code"
        },
        {
            "name": "Services",
            "model": ExcelRowService,
            "db_model": Service,
            "entity_type": "service",
            "parent_field": "pole_code"
        },
        {
            "name": "UF",
            "model": ExcelRowUF,
            "db_model": UniteFonctionnelle,
            "entity_type": "uf",
            "parent_field": "service_code"
        },
        {
            "name": "UH",
            "model": ExcelRowUH,
            "db_model": UniteHebergement,
            "entity_type": "uh",
            "parent_field": "service_code"
        },
        {
            "name": "Chambres",
            "model": ExcelRowChambre,
            "db_model": Chambre,
            "entity_type": "chambre",
            "parent_field": "uh_code"
        },
        {
            "name": "Lits",
            "model": ExcelRowLit,
            "db_model": Lit,
            "entity_type": "lit",
            "parent_field": "chambre_code"
        }
    ]
    
    # Cache pour les codes existants (évite requêtes multiples)
    existing_codes_cache = {}
    for config in sheet_configs:
        stmt = select(config["db_model"])
        results = session.exec(stmt).all()
        existing_codes_cache[config["entity_type"]] = {
            entity.code: entity for entity in results
        }
    
    # Traitement de chaque feuille
    for config in sheet_configs:
        sheet_name = config["name"]
        
        if sheet_name not in wb.sheetnames:
            preview.warnings.append(ImportMessage(
                severity=ImportSeverity.WARNING,
                message=f"Feuille '{sheet_name}' introuvable, ignorée",
                entity_type=config["entity_type"]
            ))
            continue
        
        ws = wb[sheet_name]
        
        # Lecture des données (ligne 1 = headers, lignes 2+ = données)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Ignorer lignes vides
            if not any(row):
                continue
            
            preview.total_rows += 1
            
            # Mapping colonnes → Pydantic model
            try:
                row_dict = _parse_excel_row(config, row)
                validated_row = config["model"](**row_dict)
            except Exception as e:
                preview.errors.append(ImportMessage(
                    severity=ImportSeverity.ERROR,
                    message=f"Validation échouée: {str(e)}",
                    entity_type=config["entity_type"],
                    entity_code=row_dict.get("code", "?"),
                    row_number=row_idx
                ))
                continue
            
            # Vérifier si l'entité existe déjà
            existing_entity = existing_codes_cache[config["entity_type"]].get(validated_row.code)
            
            # Déterminer l'action selon le mode
            action = _determine_action(mode, existing_entity)
            
            # Validation des références parent
            parent_code = getattr(validated_row, config["parent_field"], None) if config["parent_field"] else None
            parent_valid = True
            
            if parent_code:
                parent_type = _get_parent_type(config["entity_type"])
                if parent_type and parent_code not in existing_codes_cache[parent_type]:
                    preview.errors.append(ImportMessage(
                        severity=ImportSeverity.ERROR,
                        message=f"Parent '{parent_code}' introuvable pour {validated_row.code}",
                        entity_type=config["entity_type"],
                        entity_code=validated_row.code,
                        row_number=row_idx
                    ))
                    parent_valid = False
                    action = ImportAction.ERROR
            
            # Créer l'aperçu de l'entité
            entity_preview = ImportEntityPreview(
                entity_type=config["entity_type"],
                code=validated_row.code,
                nom=validated_row.nom,
                action=action,
                parent_code=parent_code,
                row_number=row_idx
            )
            
            # Ajouter aux listes selon l'action
            if action == ImportAction.CREATE:
                preview.to_create.append(entity_preview)
            elif action == ImportAction.UPDATE:
                preview.to_update.append(entity_preview)
            elif action == ImportAction.SKIP:
                preview.to_skip.append(entity_preview)
            elif action == ImportAction.ERROR:
                if parent_valid:  # Si erreur non liée au parent
                    preview.errors.append(ImportMessage(
                        severity=ImportSeverity.ERROR,
                        message=f"Action invalide pour {validated_row.code}",
                        entity_type=config["entity_type"],
                        entity_code=validated_row.code,
                        row_number=row_idx
                    ))
    
    duration = time.time() - start_time
    
    return {
        **preview.dict(),
        "duration_seconds": round(duration, 2),
        "can_proceed": preview.can_proceed
    }


@router.post("/import/confirm")
async def import_excel_confirm(
    file: UploadFile = File(...),
    mode: Literal["create", "update", "replace"] = Form(...),
    session: Session = Depends(get_session)
):
    """
    Exécute l'import après validation (transactionnel avec rollback)
    """
    from openpyxl import load_workbook
    from app.schemas.import_schemas import (
        ImportMode, ImportResult, ImportMessage, ImportSeverity,
        ExcelRowEG, ExcelRowPole, ExcelRowService, ExcelRowUF,
        ExcelRowUH, ExcelRowChambre, ExcelRowLit
    )
    import time
    
    start_time = time.time()
    
    # Relecture du fichier
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Erreur lecture Excel: {str(e)}")
    
    result = ImportResult(success=False)
    
    # Configuration des modèles
    sheet_configs = [
        {"name": "EntitesGeographiques", "model": ExcelRowEG, "db_model": EntiteGeographique, "entity_type": "eg"},
        {"name": "Poles", "model": ExcelRowPole, "db_model": Pole, "entity_type": "pole"},
        {"name": "Services", "model": ExcelRowService, "db_model": Service, "entity_type": "service"},
        {"name": "UF", "model": ExcelRowUF, "db_model": UniteFonctionnelle, "entity_type": "uf"},
        {"name": "UH", "model": ExcelRowUH, "db_model": UniteHebergement, "entity_type": "uh"},
        {"name": "Chambres", "model": ExcelRowChambre, "db_model": Chambre, "entity_type": "chambre"},
        {"name": "Lits", "model": ExcelRowLit, "db_model": Lit, "entity_type": "lit"}
    ]
    
    try:
        # Début transaction
        session.begin_nested()
        
        # Traitement séquentiel (respecte ordre hiérarchique)
        for config in sheet_configs:
            if config["name"] not in wb.sheetnames:
                continue
            
            ws = wb[config["name"]]
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                try:
                    row_dict = _parse_excel_row(config, row)
                    validated_row = config["model"](**row_dict)
                    
                    # Chercher entité existante
                    stmt = select(config["db_model"]).where(config["db_model"].code == validated_row.code)
                    existing = session.exec(stmt).first()
                    
                    if mode == "create" and existing:
                        result.skipped_count += 1
                        continue
                    
                    if mode == "update" and not existing:
                        result.skipped_count += 1
                        continue
                    
                    # Résoudre les foreign keys
                    fk_values = _resolve_foreign_keys(config, validated_row, session)
                    
                    if existing and mode in ["update", "replace"]:
                        # Mise à jour
                        for key, value in validated_row.dict(exclude_unset=True).items():
                            if key not in ["code"]:  # Ne pas modifier le code
                                setattr(existing, key, value)
                        for fk_key, fk_value in fk_values.items():
                            setattr(existing, fk_key, fk_value)
                        result.updated_count += 1
                    else:
                        # Création
                        entity_data = validated_row.dict(exclude_unset=True)
                        entity_data.update(fk_values)
                        new_entity = config["db_model"](**entity_data)
                        session.add(new_entity)
                        result.created_count += 1
                
                except Exception as e:
                    result.error_count += 1
                    result.messages.append(ImportMessage(
                        severity=ImportSeverity.ERROR,
                        message=f"Erreur ligne {row_idx}: {str(e)}",
                        entity_type=config["entity_type"],
                        row_number=row_idx
                    ))
        
        # Commit si aucune erreur critique
        if result.error_count == 0:
            session.commit()
            result.success = True
            result.messages.append(ImportMessage(
                severity=ImportSeverity.INFO,
                message=f"Import réussi: {result.created_count} créés, {result.updated_count} mis à jour",
                entity_type="all"
            ))
        else:
            session.rollback()
            result.success = False
            result.messages.append(ImportMessage(
                severity=ImportSeverity.ERROR,
                message=f"Import annulé suite aux {result.error_count} erreurs",
                entity_type="all"
            ))
    
    except Exception as e:
        session.rollback()
        result.success = False
        result.messages.append(ImportMessage(
            severity=ImportSeverity.ERROR,
            message=f"Erreur fatale: {str(e)}",
            entity_type="all"
        ))
    
    result.duration_seconds = round(time.time() - start_time, 2)
    
    return result.dict()


# === Helper Functions ===

def _parse_excel_row(config, row):
    """Parse une ligne Excel selon le type d'entité"""
    entity_type = config["entity_type"]
    
    if entity_type == "eg":
        return {
            "code": str(row[0]) if row[0] else "",
            "nom": str(row[1]) if row[1] else "",
            "description": str(row[2]) if row[2] else None,
            "adresse": str(row[3]) if row[3] else None,
            "ville": str(row[4]) if row[4] else None,
            "code_postal": str(row[5]) if row[5] else None,
            "telephone": str(row[6]) if row[6] else None
        }
    elif entity_type == "pole":
        return {
            "code": str(row[0]) if row[0] else "",
            "nom": str(row[1]) if row[1] else "",
            "eg_code": str(row[2]) if row[2] else "",
            "description": str(row[3]) if row[3] else None,
            "responsable": str(row[4]) if row[4] else None
        }
    elif entity_type == "service":
        return {
            "code": str(row[0]) if row[0] else "",
            "nom": str(row[1]) if row[1] else "",
            "pole_code": str(row[2]) if row[2] else "",
            "type_service": str(row[3]) if row[3] else None,
            "description": str(row[4]) if row[4] else None,
            "responsable": str(row[5]) if row[5] else None
        }
    elif entity_type == "uf":
        return {
            "code": str(row[0]) if row[0] else "",
            "nom": str(row[1]) if row[1] else "",
            "service_code": str(row[2]) if row[2] else "",
            "type_activite": str(row[3]) if row[3] else None,
            "capacite": int(row[4]) if row[4] and str(row[4]).isdigit() else None,
            "description": str(row[5]) if row[5] else None
        }
    elif entity_type == "uh":
        return {
            "code": str(row[0]) if row[0] else "",
            "nom": str(row[1]) if row[1] else "",
            "service_code": str(row[2]) if row[2] else "",
            "capacite": int(row[3]) if row[3] and str(row[3]).isdigit() else None,
            "etage": str(row[4]) if row[4] else None
        }
    elif entity_type == "chambre":
        return {
            "code": str(row[0]) if row[0] else "",
            "nom": str(row[1]) if row[1] else "",
            "uh_code": str(row[2]) if row[2] else "",
            "capacite": int(row[3]) if row[3] and str(row[3]).isdigit() else None,
            "numero": str(row[4]) if row[4] else None
        }
    elif entity_type == "lit":
        return {
            "code": str(row[0]) if row[0] else "",
            "nom": str(row[1]) if row[1] else "",
            "chambre_code": str(row[2]) if row[2] else "",
            "numero": str(row[3]) if row[3] else None
        }
    
    return {}


def _determine_action(mode: str, existing_entity) -> ImportAction:
    """Détermine l'action à effectuer selon le mode et l'existence"""
    if mode == "create":
        return ImportAction.CREATE if not existing_entity else ImportAction.SKIP
    elif mode == "update":
        return ImportAction.UPDATE if existing_entity else ImportAction.SKIP
    elif mode == "replace":
        return ImportAction.UPDATE if existing_entity else ImportAction.CREATE
    return ImportAction.ERROR


def _get_parent_type(entity_type: str) -> Optional[str]:
    """Retourne le type de l'entité parente"""
    mapping = {
        "pole": "eg",
        "service": "pole",
        "uf": "service",
        "uh": "service",
        "chambre": "uh",
        "lit": "chambre"
    }
    return mapping.get(entity_type)


def _resolve_foreign_keys(config, validated_row, session: Session) -> dict:
    """Résout les codes parents en IDs pour les foreign keys"""
    fk_values = {}
    entity_type = config["entity_type"]
    
    if entity_type == "pole" and hasattr(validated_row, "eg_code"):
        stmt = select(EntiteGeographique).where(EntiteGeographique.code == validated_row.eg_code)
        eg = session.exec(stmt).first()
        if eg:
            fk_values["entite_geographique_id"] = eg.id
    
    elif entity_type == "service" and hasattr(validated_row, "pole_code"):
        stmt = select(Pole).where(Pole.code == validated_row.pole_code)
        pole = session.exec(stmt).first()
        if pole:
            fk_values["pole_id"] = pole.id
    
    elif entity_type == "uf" and hasattr(validated_row, "service_code"):
        stmt = select(Service).where(Service.code == validated_row.service_code)
        service = session.exec(stmt).first()
        if service:
            fk_values["service_id"] = service.id
    
    elif entity_type == "uh" and hasattr(validated_row, "service_code"):
        stmt = select(Service).where(Service.code == validated_row.service_code)
        service = session.exec(stmt).first()
        if service:
            fk_values["service_id"] = service.id
    
    elif entity_type == "chambre" and hasattr(validated_row, "uh_code"):
        stmt = select(UniteHebergement).where(UniteHebergement.code == validated_row.uh_code)
        uh = session.exec(stmt).first()
        if uh:
            fk_values["unite_hebergement_id"] = uh.id
    
    elif entity_type == "lit" and hasattr(validated_row, "chambre_code"):
        stmt = select(Chambre).where(Chambre.code == validated_row.chambre_code)
        chambre = session.exec(stmt).first()
        if chambre:
            fk_values["chambre_id"] = chambre.id
    
    return fk_values


# === UI Routes ===

@ui_router.get("/import", response_class=HTMLResponse)
async def structure_import_page(request: Request):
    """Page d'import de structure"""
    return templates.TemplateResponse(
        "structure_import.html",
        {"request": request}
    )
